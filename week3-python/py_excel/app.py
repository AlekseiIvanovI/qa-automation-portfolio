import openpyxl

# wb = openpyxl.Workbook()
wb = openpyxl.load_workbook("transactions.xlsx")

print(wb.sheetnames)
sheet = wb["Sheet1"]

cell = sheet["a1"]
column = sheet["a"]
cells = sheet["a:c"]
print(cells)
print(column)
sheet.cell(row=1, column=1)

print(sheet.max_row)
print(sheet.max_column)


for row in range(1, sheet.max_row + 1):
    for column in range(1, sheet.max_column + 1):
        cell = sheet.cell(row, column)
        print(cell.value)
sheet.append([1, 2, 3])
wb.save("transactions2.xlsx")
# print(cell.value)
# print(cell.row)
# print(cell.column)
# print(cell.coordinate)
